#!/usr/bin/env python3
"""Code the TECHNOLOGY field of the FDI dataset into a specific component + annual capacity.

This is the mechanical half of the `code-fdi-components` skill. Claude does the
judgment (reads each row's source URLs, classifies the component, extracts the
annual production capacity) and writes a small batch JSON; this script backs up
the workbook, validates the batch against the admissible-component taxonomy, and
merges it into FDI_All_ToSource.xlsx by Project ID. It never invents a coding.

Subcommands
-----------
  init       Back up the workbook and add the 7 coding columns (idempotent).
  apply      Validate a coding batch JSON and merge it into the workbook by Project ID.
  validate   Validate every coded row in the workbook; print coverage stats.

Exit codes: 0 = clean, 1 = validation violation(s), 2 = usage / IO error.

Coding columns (appended in place; original columns untouched)
  Component | Component HS6 | Capacity Value | Capacity Unit | Capacity Source | Coding Tier | Coding Notes

Batch JSON shape (array of objects)
  [
    {
      "project_id": "HAR0010",
      "technology": "Battery",            # optional, disambiguates duplicate Project IDs
      "component": "Lithium-ion battery (cell/pack)",
      "component_hs6": "850760",
      "capacity_value": 100,
      "capacity_unit": "GWh/year",
      "capacity_source": "https://www.catl.com/en/news/983.html",
      "coding_tier": 1,
      "coding_notes": "Cells + modules; modules secondary. EUR7.34B is capex, not capacity."
    }
  ]

stdlib + openpyxl only.
"""

from __future__ import annotations

import argparse
import json
import shutil
import sys
from datetime import datetime
from pathlib import Path
from urllib.parse import urlparse

try:
    import openpyxl
except ImportError:  # pragma: no cover
    print("error: openpyxl required (pip install openpyxl)", file=sys.stderr)
    raise SystemExit(2)

# ---------------------------------------------------------------------------
# Layout
# ---------------------------------------------------------------------------

SCRIPT_DIR = Path(__file__).resolve().parent
TAXONOMY_PATH = SCRIPT_DIR.parent / "references" / "component_taxonomy.json"
WORKBOOK_NAME = "FDI_All_ToSource.xlsx"
# Default workbook: two levels up from data/ (projects/nzipl/FDI_All_ToSource.xlsx).
DEFAULT_XLSX = SCRIPT_DIR.parent.parent / WORKBOOK_NAME


def find_workbook(explicit):
    """Resolve the target workbook so the skill runs from any checkout location.

    Order: explicit --xlsx; the relative default next to the skill; then an
    upward search (up to 6 levels) from the cwd and the script dir for
    FDI_All_ToSource.xlsx. Falls back to the relative default so the caller's
    existence check produces a clear error.
    """
    if explicit:
        return Path(explicit)
    if DEFAULT_XLSX.exists():
        return DEFAULT_XLSX
    for base in (Path.cwd(), SCRIPT_DIR):
        p = base.resolve()
        for _ in range(6):
            cand = p / WORKBOOK_NAME
            if cand.exists():
                return cand
            if p.parent == p:
                break
            p = p.parent
    return DEFAULT_XLSX

SHEET = "All"
TECH_COL = "Technology"
ID_COL = "Project ID"

CODING_COLUMNS = [
    "Component",
    "Component HS6",
    "Capacity Value",
    "Capacity Unit",
    "Capacity Source",
    "Coding Tier",
    "Coding Notes",
]

TECHS = {"Solar", "Wind", "Battery", "EV"}

FORBIDDEN_DOMAINS = {
    "wikipedia.org", "en.wikipedia.org", "grokipedia.com",
    "linkedin.com", "www.linkedin.com", "reddit.com", "www.reddit.com",
    "twitter.com", "x.com", "facebook.com", "www.facebook.com",
    "instagram.com", "www.instagram.com", "medium.com",
}

# Finished-cell components whose canonical unit is energy (GWh), used by the
# GW-vs-GWh sanity heuristic.
FINISHED_CELL_COMPONENTS = {
    "Lithium-ion battery (cell/pack)",
    "Ni-MH battery (cell/pack)",
    "Battery pack / module assembly",
}


# ---------------------------------------------------------------------------
# Taxonomy
# ---------------------------------------------------------------------------

def load_taxonomy(path: Path = TAXONOMY_PATH) -> dict:
    with open(path) as f:
        return json.load(f)


def norm_tech(value) -> str:
    """Normalize the Technology cell (trailing spaces appear in the raw data)."""
    return str(value).strip() if value is not None else ""


def domain_of(url: str) -> str:
    try:
        host = urlparse(url).netloc.lower()
        return host[4:] if host.startswith("www.") else host
    except Exception:
        return ""


# ---------------------------------------------------------------------------
# Validation (shared by apply + validate)
# ---------------------------------------------------------------------------

def validate_coding(row: dict, tax: dict) -> tuple[list[str], list[str]]:
    """Validate one coded row. Returns (errors, warnings).

    `row` keys (any may be None/blank): project_id, technology, component,
    component_hs6, capacity_value, capacity_unit, capacity_source, coding_tier.
    Only rows with a non-empty `component` are validated; blank rows are skipped
    by the caller.
    """
    errs: list[str] = []
    warns: list[str] = []

    rid = row.get("project_id") or "<no-id>"
    tech = norm_tech(row.get("technology"))
    comp = (row.get("component") or "").strip()
    hs6 = (str(row.get("component_hs6")).strip() if row.get("component_hs6") not in (None, "") else "")
    cap_val = row.get("capacity_value")
    cap_unit = (row.get("capacity_unit") or "").strip()
    cap_src = (row.get("capacity_source") or "").strip()
    tier = row.get("coding_tier")

    label = tax["out_of_shortlist_label"]
    unit_enum = set(tax["capacity_unit_enum"])

    if tech not in TECHS:
        errs.append(f"{rid}: technology '{tech}' not one of {sorted(TECHS)}")
        return errs, warns  # can't validate component without a known tech

    comps = tax["technologies"][tech]["components"]

    # --- Component ---
    is_other = comp == label
    known = comp in comps
    if not comp:
        errs.append(f"{rid}: Component is blank")
    elif not known and not is_other:
        errs.append(
            f"{rid}: Component '{comp}' is not an admissible {tech} component "
            f"and is not '{label}'"
        )

    # --- HS6 ---
    if known:
        allowed_codes = {comps[comp]["hs6"], comps[comp]["hs6_shortlist_key"]}
        if not hs6:
            warns.append(f"{rid}: Component HS6 missing for '{comp}' (expected one of {sorted(allowed_codes)})")
        elif not (hs6.isdigit() and len(hs6) == 6):
            errs.append(f"{rid}: Component HS6 '{hs6}' must be a 6-digit code")
        elif hs6 not in allowed_codes:
            errs.append(
                f"{rid}: Component HS6 '{hs6}' not valid for '{comp}' "
                f"(expected customs {comps[comp]['hs6']} or shortlist-key {comps[comp]['hs6_shortlist_key']})"
            )
    elif is_other:
        if hs6 and not (hs6.isdigit() and len(hs6) == 6):
            errs.append(f"{rid}: Component HS6 '{hs6}' must be a 6-digit code (or blank for out-of-shortlist)")
        if not (row.get("coding_notes") or "").strip():
            warns.append(f"{rid}: out-of-shortlist row should explain in Coding Notes what it actually makes")

    # --- Capacity value/unit pairing ---
    has_val = cap_val not in (None, "")
    has_unit = bool(cap_unit)
    if has_val:
        try:
            v = float(cap_val)
            if v <= 0:
                errs.append(f"{rid}: Capacity Value must be positive (got {cap_val})")
        except (TypeError, ValueError):
            errs.append(f"{rid}: Capacity Value '{cap_val}' is not numeric")
        if not has_unit:
            errs.append(f"{rid}: Capacity Value present without Capacity Unit")
    if has_unit:
        if not has_val:
            errs.append(f"{rid}: Capacity Unit present without Capacity Value")
        if cap_unit not in unit_enum:
            errs.append(f"{rid}: Capacity Unit '{cap_unit}' not in the unit enum")
        elif known and cap_unit not in comps[comp]["capacity_units"]:
            warns.append(
                f"{rid}: Capacity Unit '{cap_unit}' is unusual for '{comp}' "
                f"(typical: {comps[comp]['capacity_units']})"
            )

    # --- GW vs GWh heuristic ---
    if cap_unit == "GWh/year" and tech in {"Solar", "Wind"}:
        warns.append(f"{rid}: GWh on a {tech} row almost always means a misfiled battery plant — re-check the Technology tag")
    if cap_unit == "GWh/year" and tech == "EV":
        warns.append(f"{rid}: GWh on an EV row — this is a battery plant; it should be Technology=Battery, not EV")
    if has_unit and comp in FINISHED_CELL_COMPONENTS and cap_unit not in {"GWh/year", "packs/year", "MWh/year"}:
        warns.append(f"{rid}: finished cell/pack capacity is normally GWh/year, not '{cap_unit}'")

    # --- Capacity source ---
    if cap_src:
        if not cap_src.startswith(("http://", "https://")):
            errs.append(f"{rid}: Capacity Source must be an http(s) URL (got '{cap_src[:50]}')")
        elif domain_of(cap_src) in FORBIDDEN_DOMAINS:
            errs.append(f"{rid}: Capacity Source uses forbidden domain '{domain_of(cap_src)}'")

    # --- Tier ---
    if tier in (None, ""):
        if comp:
            warns.append(f"{rid}: Coding Tier missing")
    else:
        try:
            t = int(tier)
        except (TypeError, ValueError):
            errs.append(f"{rid}: Coding Tier '{tier}' must be 1, 2, or 3")
            t = None
        if t is not None:
            if t not in (1, 2, 3):
                errs.append(f"{rid}: Coding Tier {t} must be 1, 2, or 3")
            elif t == 1 and not (has_val and has_unit and cap_src):
                errs.append(f"{rid}: Tier 1 requires Capacity Value + Unit + Source (component and capacity both sourced)")

    return errs, warns


# ---------------------------------------------------------------------------
# Workbook helpers
# ---------------------------------------------------------------------------

def backup(xlsx: Path) -> Path:
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    bdir = xlsx.parent / "backups"
    bdir.mkdir(exist_ok=True)
    dest = bdir / f"{xlsx.stem}.{stamp}{xlsx.suffix}"
    shutil.copy2(xlsx, dest)
    return dest


def header_map(ws) -> dict:
    return {c.value: i for i, c in enumerate(next(ws.iter_rows(min_row=1, max_row=1))) if c.value is not None}


# ---------------------------------------------------------------------------
# init
# ---------------------------------------------------------------------------

def cmd_init(args) -> int:
    xlsx = find_workbook(args.xlsx)
    if not xlsx.exists():
        print(f"error: {xlsx} does not exist", file=sys.stderr)
        return 2
    b = backup(xlsx)
    print(f"backup: {b}")

    wb = openpyxl.load_workbook(xlsx)
    ws = wb[SHEET] if SHEET in wb.sheetnames else wb.active
    hdr = header_map(ws)
    existing = set(hdr)
    next_col = ws.max_column
    added = []
    for name in CODING_COLUMNS:
        if name in existing:
            continue
        next_col += 1
        ws.cell(row=1, column=next_col, value=name)
        added.append(name)
    if added:
        wb.save(xlsx)
        print(f"added columns: {', '.join(added)}")
    else:
        print("all coding columns already present; no change")
    return 0


# ---------------------------------------------------------------------------
# apply
# ---------------------------------------------------------------------------

def cmd_apply(args) -> int:
    xlsx = find_workbook(args.xlsx)
    batch_path = Path(args.batch)
    if not xlsx.exists():
        print(f"error: {xlsx} does not exist", file=sys.stderr)
        return 2
    if not batch_path.exists():
        print(f"error: {batch_path} does not exist", file=sys.stderr)
        return 2

    tax = load_taxonomy()
    with open(batch_path) as f:
        batch = json.load(f)
    if not isinstance(batch, list):
        print("error: batch JSON must be an array of coding objects", file=sys.stderr)
        return 2

    # Validate the whole batch first; refuse to write if anything fails.
    all_errs: list[str] = []
    all_warns: list[str] = []
    for row in batch:
        e, w = validate_coding(row, tax)
        all_errs.extend(e)
        all_warns.extend(w)
    for w in all_warns:
        print(f"  warn: {w}")
    if all_errs:
        print(f"FAIL: {len(all_errs)} violation(s) in batch — nothing written")
        for e in all_errs:
            print(f"  - {e}")
        return 1

    wb = openpyxl.load_workbook(xlsx)
    ws = wb[SHEET] if SHEET in wb.sheetnames else wb.active
    hdr = header_map(ws)
    missing_cols = [c for c in CODING_COLUMNS if c not in hdr]
    if missing_cols:
        print(f"error: workbook is missing coding columns {missing_cols}; run `init` first", file=sys.stderr)
        return 2
    if ID_COL not in hdr or TECH_COL not in hdr:
        print(f"error: workbook missing '{ID_COL}' or '{TECH_COL}' column", file=sys.stderr)
        return 2

    # Index rows by Project ID (with Technology for disambiguation).
    rows_by_id: dict[str, list[int]] = {}
    for ridx in range(2, ws.max_row + 1):
        pid = ws.cell(row=ridx, column=hdr[ID_COL] + 1).value
        if pid is None:
            continue
        rows_by_id.setdefault(str(pid).strip(), []).append(ridx)

    b = backup(xlsx)
    print(f"backup: {b}")

    written = 0
    skipped: list[str] = []
    for row in batch:
        pid = str(row.get("project_id") or "").strip()
        candidates = rows_by_id.get(pid, [])
        if not candidates:
            skipped.append(f"{pid}: no matching row")
            continue
        if len(candidates) > 1:
            want_tech = norm_tech(row.get("technology"))
            candidates = [
                ri for ri in candidates
                if norm_tech(ws.cell(row=ri, column=hdr[TECH_COL] + 1).value) == want_tech
            ]
            if len(candidates) != 1:
                skipped.append(f"{pid}: {len(candidates)} rows match (provide a disambiguating 'technology')")
                continue
        ridx = candidates[0]
        _set(ws, ridx, hdr, "Component", row.get("component"))
        _set(ws, ridx, hdr, "Component HS6", row.get("component_hs6"))
        _set(ws, ridx, hdr, "Capacity Value", row.get("capacity_value"))
        _set(ws, ridx, hdr, "Capacity Unit", row.get("capacity_unit"))
        _set(ws, ridx, hdr, "Capacity Source", row.get("capacity_source"))
        _set(ws, ridx, hdr, "Coding Tier", row.get("coding_tier"))
        _set(ws, ridx, hdr, "Coding Notes", row.get("coding_notes"))
        written += 1

    wb.save(xlsx)
    print(f"OK: wrote {written} row(s)")
    if skipped:
        print(f"skipped {len(skipped)}:")
        for s in skipped:
            print(f"  - {s}")
    return 0


def _set(ws, ridx: int, hdr: dict, col: str, value):
    if value in (None, ""):
        return
    ws.cell(row=ridx, column=hdr[col] + 1, value=value)


# ---------------------------------------------------------------------------
# validate
# ---------------------------------------------------------------------------

def cmd_validate(args) -> int:
    xlsx = find_workbook(args.xlsx)
    if not xlsx.exists():
        print(f"error: {xlsx} does not exist", file=sys.stderr)
        return 2
    tax = load_taxonomy()
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb[SHEET] if SHEET in wb.sheetnames else wb.active
    it = ws.iter_rows(values_only=True)
    hdr = {h: i for i, h in enumerate(next(it)) if h is not None}
    for c in CODING_COLUMNS + [ID_COL, TECH_COL]:
        if c not in hdr:
            print(f"error: workbook missing column '{c}'; run `init` first", file=sys.stderr)
            return 2

    def g(r, col):
        return r[hdr[col]] if hdr[col] < len(r) else None

    rows = list(it)
    all_errs: list[str] = []
    all_warns: list[str] = []

    # coverage accumulators
    total = 0
    coded = 0
    per_tech: dict[str, dict] = {t: {"total": 0, "coded": 0, "with_cap": 0, "other": 0} for t in TECHS}
    tier_dist = {1: 0, 2: 0, 3: 0}

    for r in rows:
        total += 1
        tech = norm_tech(g(r, TECH_COL))
        if tech in per_tech:
            per_tech[tech]["total"] += 1
        comp = g(r, "Component")
        if comp in (None, ""):
            continue  # uncoded row — skip validation, counts as not coded
        coded += 1
        row = {
            "project_id": g(r, ID_COL),
            "technology": tech,
            "component": comp,
            "component_hs6": g(r, "Component HS6"),
            "capacity_value": g(r, "Capacity Value"),
            "capacity_unit": g(r, "Capacity Unit"),
            "capacity_source": g(r, "Capacity Source"),
            "coding_tier": g(r, "Coding Tier"),
            "coding_notes": g(r, "Coding Notes"),
        }
        e, w = validate_coding(row, tax)
        all_errs.extend(e)
        all_warns.extend(w)
        if tech in per_tech:
            per_tech[tech]["coded"] += 1
            if g(r, "Capacity Value") not in (None, ""):
                per_tech[tech]["with_cap"] += 1
            if comp == tax["out_of_shortlist_label"]:
                per_tech[tech]["other"] += 1
        try:
            tier_dist[int(g(r, "Coding Tier"))] += 1
        except (TypeError, ValueError):
            pass

    # report
    print(f"=== coverage ({coded}/{total} rows coded) ===")
    for t in ("Solar", "Wind", "Battery", "EV"):
        d = per_tech[t]
        if d["total"]:
            print(f"  {t:8s} {d['coded']:3d}/{d['total']:3d} coded | {d['with_cap']:3d} with capacity | {d['other']:3d} out-of-shortlist")
    print(f"  tiers: 1={tier_dist[1]}  2={tier_dist[2]}  3={tier_dist[3]}")

    if all_warns and not args.quiet_warnings:
        print(f"--- {len(all_warns)} warning(s) ---")
        for w in all_warns:
            print(f"  warn: {w}")

    if all_errs:
        print(f"FAIL: {len(all_errs)} violation(s)")
        for e in all_errs:
            print(f"  - {e}")
        return 1
    print(f"OK: {coded} coded row(s) clean")
    return 0


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main() -> int:
    p = argparse.ArgumentParser(description="Code FDI TECHNOLOGY into component + annual capacity.")
    sub = p.add_subparsers(dest="cmd", required=True)

    pi = sub.add_parser("init", help="back up the workbook and add the 7 coding columns")
    pi.add_argument("--xlsx", default=None, help="workbook path (default: auto-detected FDI_All_ToSource.xlsx)")
    pi.set_defaults(func=cmd_init)

    pa = sub.add_parser("apply", help="validate a coding batch JSON and merge it into the workbook")
    pa.add_argument("batch", help="path to the batch JSON (array of coding objects)")
    pa.add_argument("--xlsx", default=None, help="workbook path (default: auto-detected FDI_All_ToSource.xlsx)")
    pa.set_defaults(func=cmd_apply)

    pv = sub.add_parser("validate", help="validate every coded row; print coverage stats")
    pv.add_argument("--xlsx", default=None, help="workbook path (default: auto-detected FDI_All_ToSource.xlsx)")
    pv.add_argument("--quiet-warnings", action="store_true", help="suppress per-row warnings")
    pv.set_defaults(func=cmd_validate)

    args = p.parse_args()
    return args.func(args)


if __name__ == "__main__":
    raise SystemExit(main())
