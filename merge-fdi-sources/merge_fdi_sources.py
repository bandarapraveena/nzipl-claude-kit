"""Merge a curated FDI source list into FDI_All_ToSource.xlsx.

Auto-detects the input format from the workbook layout:

  * China list  (e.g. China_FDI_Battery.xlsx, sheet = tech name)
      - Match key: HAR* IDs in column 1
      - Source cols 15-22, Capital Investment col 11
      - Empty marker: '\\'

  * Global manual list  (e.g. Globa_FDI_Battery_Manual.xlsx, single sheet)
      - Match key: FDI* IDs in column 1
      - Source cols 18,19,20,21,22,23,24,26 (skips 25 = Tech source)
      - Capital Investment col 13
      - Empty marker: 'NA', '-'

Precedence rules (identical for both inputs):
  1. Column S (Source for Amount) is the priority cell.
  2. Existing master URL beats curated URL — including '[UNVERIFIED] ...'.
     Only cells equal to 'fDi Markets (FT)' are replaceable.
  3. If S is replaced and the curated row has a real Capital Investment
     value, update master L too.
  4. Empty master cells are not replaced (this skill does not invent sources).

Yellow PatternFill (FFFFFF00) is applied to every changed cell.
"""
import argparse
import sys
import openpyxl
from openpyxl.styles import PatternFill

YELLOW = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
FALLBACK = "fDi Markets (FT)"
DISCLOSURE_NOTE = "Investment value not publicly disclosed"

MASTER_SOURCE_COLS = {"P": 16, "Q": 17, "R": 18, "S": 19, "T": 20, "U": 21, "V": 22, "W": 23}
MASTER_Q_COL = 17
MASTER_S_COL = 19
MASTER_L_COL = 12
MASTER_ID_COL = 2
MASTER_SECTOR_COL = 1

LAYOUTS = {
    "china": {
        "id_prefix": "HAR",
        "id_col": 1,
        "source_cols": {"P": 15, "Q": 16, "R": 17, "S": 18, "T": 19, "U": 20, "V": 21, "W": 22},
        "l_col": 11,
        "empty_markers": {"\\"},
    },
    "manual": {
        "id_prefix": "FDI",
        "id_col": 1,
        "source_cols": {"P": 18, "Q": 19, "R": 20, "S": 21, "T": 22, "U": 23, "V": 24, "W": 26},
        "l_col": 13,
        "empty_markers": {"NA", "-"},
    },
}


def detect_layout(ws, hint=None):
    """Return ('china'|'manual', layout_dict). Hint overrides auto-detect."""
    if hint in LAYOUTS:
        return hint, LAYOUTS[hint]
    har = fdi = 0
    for r in range(2, min(ws.max_row, 200) + 1):
        v = ws.cell(r, 1).value
        if not v:
            continue
        s = str(v).strip()
        if s.startswith("HAR"):
            har += 1
        elif s.startswith("FDI"):
            fdi += 1
    if har > fdi:
        return "china", LAYOUTS["china"]
    if fdi > 0:
        return "manual", LAYOUTS["manual"]
    sys.exit("Could not auto-detect input layout: no HAR* or FDI* IDs in column 1.")


def is_real_url(v, empty_markers):
    if v is None:
        return False
    s = str(v).strip()
    if not s or s == FALLBACK or s == DISCLOSURE_NOTE:
        return False
    if any(s.upper() == m.upper() for m in empty_markers):
        return False
    return True


def to_number(v, empty_markers):
    if v is None:
        return None
    s = str(v).strip()
    if not s or any(s.upper() == m.upper() for m in empty_markers):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def is_credible_url(v):
    """A 'credible' source cell — anything that's a real URL (incl. [UNVERIFIED])."""
    if v is None:
        return False
    s = str(v).strip()
    if not s or s == FALLBACK or s == DISCLOSURE_NOTE:
        return False
    return True


def s_is_first_best(v):
    """First-best = column S has a credible URL (a real source for the amount)."""
    return is_credible_url(v)


def s_needs_disclosure(v):
    """S is FT or empty — eligible for the disclosure note when Q is credible."""
    if v is None:
        return True
    s = str(v).strip()
    return s == "" or s == FALLBACK


def stage2_pass(ws_m, ws_c, layout, sector_filter):
    """Second-stage verification — runs after the standard stage-1 merge.

    Two sub-steps, scoped to `sector_filter` if given:
      (a) Relaxed Q fill: for rows that have NOT met first-best (S not credible),
          fill Q from the curated file if master Q is empty or FT. This relaxes
          stage-1 Rule 6 (which leaves empty cells alone) for col Q only.
      (b) Disclosure note: for rows where Q is now credible and S is still
          FT/empty, set S to 'Investment value not publicly disclosed'.
    """
    empty = layout["empty_markers"]
    curated_q_col = layout["source_cols"]["Q"]
    id_col = layout["id_col"]

    # Stage 2 indexes by ANY ID in the curated file — sector is the filter,
    # not ID prefix. The manual file mixes FDI* and Battery* IDs; the China
    # file is HAR* only. Both should match into stage 2.
    curated_idx = {}
    for r in range(2, ws_c.max_row + 1):
        pid = ws_c.cell(r, id_col).value
        if pid:
            curated_idx[str(pid).strip()] = r

    sector_check = (lambda s: True) if not sector_filter else (
        lambda s: s is not None and sector_filter.lower() in str(s).lower()
    )

    q_filled, s_disclosed = [], []

    # (a) Relaxed Q fill
    for r in range(2, ws_m.max_row + 1):
        if not sector_check(ws_m.cell(r, MASTER_SECTOR_COL).value):
            continue
        if s_is_first_best(ws_m.cell(r, MASTER_S_COL).value):
            continue
        q_cell = ws_m.cell(r, MASTER_Q_COL)
        qv = q_cell.value
        qs = "" if qv is None else str(qv).strip()
        if qs and qs != FALLBACK:
            continue  # Q already credible
        pid = ws_m.cell(r, MASTER_ID_COL).value
        if pid is None or str(pid).strip() not in curated_idx:
            continue
        cval = ws_c.cell(curated_idx[str(pid).strip()], curated_q_col).value
        if not is_real_url(cval, empty):
            continue
        q_cell.value = cval
        q_cell.fill = YELLOW
        q_filled.append((r, str(pid).strip(), str(cval)[:60]))

    # (b) Disclosure-note pass on S — also clear L (Capital Investment).
    # The FT-derived L number is no longer credibly sourced once we mark S
    # as 'not publicly disclosed', so we drop it to keep L and S consistent.
    for r in range(2, ws_m.max_row + 1):
        if not sector_check(ws_m.cell(r, MASTER_SECTOR_COL).value):
            continue
        s_cell = ws_m.cell(r, MASTER_S_COL)
        if not s_needs_disclosure(s_cell.value):
            continue
        if not is_credible_url(ws_m.cell(r, MASTER_Q_COL).value):
            continue
        s_cell.value = DISCLOSURE_NOTE
        s_cell.fill = YELLOW
        l_cell = ws_m.cell(r, MASTER_L_COL)
        old_l = l_cell.value
        l_cleared = False
        if old_l is not None and str(old_l).strip() != "":
            l_cell.value = None
            l_cell.fill = YELLOW
            l_cleared = True
        s_disclosed.append((r, str(ws_m.cell(r, MASTER_ID_COL).value).strip(), old_l if l_cleared else None))

    return q_filled, s_disclosed


def merge(master_path, curated_path, sheet_name, layout_hint, limit, run_stage2, sector_filter):
    wb_m = openpyxl.load_workbook(master_path)
    if "All" not in wb_m.sheetnames:
        sys.exit(f"Master file missing sheet 'All': {master_path}")
    ws_m = wb_m["All"]

    wb_c = openpyxl.load_workbook(curated_path, data_only=True)
    if sheet_name and sheet_name in wb_c.sheetnames:
        ws_c = wb_c[sheet_name]
    elif sheet_name:
        sys.exit(f"Curated file missing sheet '{sheet_name}'. Available: {wb_c.sheetnames}")
    else:
        ws_c = wb_c.active

    layout_name, layout = detect_layout(ws_c, layout_hint)
    print(f"[info] curated layout: {layout_name} (sheet={ws_c.title})")

    prefix = layout["id_prefix"]
    empty = layout["empty_markers"]

    # Match by exact ID, not prefix. The China file is HAR-only by construction;
    # the manual file mixes FDI*/Battery*/Batteries* IDs and we want all of them
    # eligible. id_prefix is now informational (used for the missing-IDs report).
    curated = {}
    for r in range(2, ws_c.max_row + 1):
        pid = ws_c.cell(r, layout["id_col"]).value
        if pid:
            curated[str(pid).strip()] = r

    master_index = {}
    for r in range(2, ws_m.max_row + 1):
        pid = ws_m.cell(r, MASTER_ID_COL).value
        if pid:
            master_index.setdefault(str(pid).strip(), []).append(r)

    if limit is not None:
        # Iterate in master-row order so "first N" is reviewer-friendly.
        ordered, seen = [], set()
        for r in range(2, ws_m.max_row + 1):
            pid = ws_m.cell(r, MASTER_ID_COL).value
            if not pid:
                continue
            pid = str(pid).strip()
            if pid in curated and pid not in seen:
                ordered.append(pid)
                seen.add(pid)
        pids = ordered
    else:
        pids = sorted(curated.keys())

    log = []
    missing = []
    processed = 0

    for pid in pids:
        if limit is not None and processed >= limit:
            break
        if pid not in master_index:
            missing.append(pid)
            continue
        crow = curated[pid]
        for mrow in master_index[pid]:
            if limit is not None and processed >= limit:
                break
            # Replaceable cells: 'fDi Markets (FT)' anywhere, plus the
            # disclosure-note literal in column S only (a real URL upgrades
            # a previously-disclosed row back to first-best).
            def replaceable(letter, val):
                s = "" if val is None else str(val).strip()
                if s == FALLBACK:
                    return True
                if letter == "S" and s == DISCLOSURE_NOTE:
                    return True
                return False

            if not any(replaceable(L, ws_m.cell(mrow, c).value) for L, c in MASTER_SOURCE_COLS.items()):
                continue
            row_changes = []
            s_replaced = False
            for letter, mcol in MASTER_SOURCE_COLS.items():
                cell = ws_m.cell(mrow, mcol)
                old = "" if cell.value is None else str(cell.value).strip()
                if not replaceable(letter, cell.value):
                    continue
                cval = ws_c.cell(crow, layout["source_cols"][letter]).value
                if not is_real_url(cval, empty):
                    continue
                cell.value = cval
                cell.fill = YELLOW
                marker = "DISCLOSED" if old == DISCLOSURE_NOTE else "FT"
                row_changes.append(f"{letter}({mcol}): {marker} -> {str(cval)[:60]}")
                if letter == "S":
                    s_replaced = True

            if s_replaced:
                new_l = to_number(ws_c.cell(crow, layout["l_col"]).value, empty)
                if new_l is not None:
                    l_cell = ws_m.cell(mrow, MASTER_L_COL)
                    old_l = to_number(l_cell.value, empty)
                    if old_l != new_l:
                        l_cell.value = new_l
                        l_cell.fill = YELLOW
                        row_changes.append(f"L: {old_l} -> {new_l}")

            processed += 1
            log.append((pid, mrow, row_changes or ["(no curated URLs available for FT cells)"]))

    if run_stage2:
        q_filled, s_disclosed = stage2_pass(ws_m, ws_c, layout, sector_filter)
    else:
        q_filled, s_disclosed = [], []
    wb_m.save(master_path)

    print("\n=== Change log ===")
    if not log:
        print("  (no changes)")
    for pid, row, changes in log:
        print(f"\n{pid} (master row {row}):")
        for c in changes:
            print(f"  {c}")

    if missing:
        print(f"\n=== Curated IDs not found in master ({len(missing)}) ===")
        for pid in missing[:20]:
            print(f"  {pid}")
        if len(missing) > 20:
            print(f"  ... and {len(missing) - 20} more")

    updated = sum(1 for _, _, c in log if c[0] != "(no curated URLs available for FT cells)")
    print(f"\nStage 1: processed {processed} matched row(s); {updated} updated.")

    if run_stage2:
        scope = f"sector={sector_filter}" if sector_filter else "all sectors"
        print(f"\n=== Stage 2 ({scope}) ===")
        print(f"Q (Source for Project Status) filled from curated: {len(q_filled)}")
        for r, pid, url in q_filled:
            print(f"  row {r} {pid}: {url}")
        print(f"\nS (Source for Amount) set to disclosure note: {len(s_disclosed)}")
        for entry in s_disclosed:
            r, pid, cleared_l = entry
            tail = f" (L cleared: was {cleared_l})" if cleared_l is not None else ""
            print(f"  row {r} {pid}{tail}")


def main():
    ap = argparse.ArgumentParser(description="Merge a curated FDI source list into FDI_All_ToSource.xlsx")
    ap.add_argument("--master", required=True, help="Path to FDI_All_ToSource.xlsx")
    ap.add_argument("--curated", required=True, help="Path to China_FDI_<Tech>.xlsx OR Globa_FDI_<Tech>_Manual.xlsx")
    ap.add_argument("--sheet", default=None, help="Sheet name in the curated file (default: active sheet)")
    ap.add_argument("--layout", choices=list(LAYOUTS), default=None, help="Force layout (default: auto-detect from ID prefix)")
    ap.add_argument("--limit", type=int, default=None, help="Process first N matched rows in master-row order")
    ap.add_argument("--stage2", action="store_true",
                    help="After stage-1 merge, run second-stage verification: relaxed Q-fill + disclosure-note pass on S")
    ap.add_argument("--sector", default=None,
                    help="Sector filter for stage 2 (e.g. 'battery'). Default: all sectors")
    args = ap.parse_args()
    merge(args.master, args.curated, args.sheet, args.layout, args.limit, args.stage2, args.sector)


if __name__ == "__main__":
    main()
