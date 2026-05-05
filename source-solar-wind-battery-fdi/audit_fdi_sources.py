#!/usr/bin/env python3
"""Ground-truth audit of FDI_All_ToSource.xlsx source columns (P-W).

Reports which rows are fully sourced, partially sourced, or unsourced based
on the actual cell contents — independent of any progress log. Useful when
the progress JSON has drifted from reality.

Output:
- Summary counts (fully / partial / none)
- Sorted list of unsourced row numbers
- Sorted list of partially-sourced row numbers with column letters filled
- Optional --json for machine-readable output

Usage:
    python3 audit_fdi_sources.py ../LCT-FDI/FDI_All_ToSource.xlsx [--json]
"""

import argparse
import json
import sys

import openpyxl
from openpyxl.utils import get_column_letter

SOURCE_COLUMNS = ["P", "Q", "R", "S", "T", "U", "V", "W"]


def is_filled(value) -> bool:
    if value is None:
        return False
    text = str(value).strip()
    return bool(text)


def audit(path: str) -> dict:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    fully, partial, none = [], [], []
    partial_detail = {}
    src_indices = [(get_column_letter(i), i - 1) for i in range(16, 24)]  # P-W

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        filled_cols = [
            letter
            for letter, idx in src_indices
            if idx < len(row) and is_filled(row[idx])
        ]
        row_num = row_idx
        if len(filled_cols) == len(SOURCE_COLUMNS):
            fully.append(row_num)
        elif filled_cols:
            partial.append(row_num)
            partial_detail[row_num] = filled_cols
        else:
            none.append(row_num)

    return {
        "path": path,
        "total_data_rows": len(fully) + len(partial) + len(none),
        "fully_sourced": fully,
        "partially_sourced": partial,
        "partial_detail": partial_detail,
        "unsourced": none,
    }


def compress_ranges(nums: list[int]) -> list[str]:
    if not nums:
        return []
    nums = sorted(nums)
    out, start, prev = [], nums[0], nums[0]
    for n in nums[1:]:
        if n == prev + 1:
            prev = n
        else:
            out.append(f"{start}" if start == prev else f"{start}-{prev}")
            start = prev = n
    out.append(f"{start}" if start == prev else f"{start}-{prev}")
    return out


def main() -> int:
    p = argparse.ArgumentParser()
    p.add_argument("path")
    p.add_argument("--json", action="store_true", help="emit JSON")
    args = p.parse_args()

    result = audit(args.path)

    if args.json:
        print(json.dumps(result, indent=2))
        return 0

    full = len(result["fully_sourced"])
    part = len(result["partially_sourced"])
    none = len(result["unsourced"])
    total = result["total_data_rows"]

    print(f"file: {result['path']}")
    print(f"total data rows: {total}")
    print(f"  fully sourced  (P-W all filled): {full}")
    print(f"  partial        (some filled):    {part}")
    print(f"  unsourced      (P-W all empty):  {none}")
    print()
    print("unsourced row ranges:")
    for r in compress_ranges(result["unsourced"]):
        print(f"  {r}")
    print()
    print("partially-sourced rows (row: filled cols):")
    for row in sorted(result["partial_detail"]):
        cols = "".join(result["partial_detail"][row])
        print(f"  {row}: {cols}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
