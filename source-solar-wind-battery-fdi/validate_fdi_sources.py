#!/usr/bin/env python3
"""Validate FDI source citations in columns P-W of an xlsx.

Hard-rejects forbidden-domain URLs (wikipedia, social, medium) and reports
malformed cells. Exit code 0 = clean; non-zero = errors found.

Usage:
    python3 validate_fdi_sources.py FDI_Combined.xlsx
"""

import re
import sys
from urllib.parse import urlparse

import openpyxl

FORBIDDEN_HOSTS = {
    "wikipedia.org",
    "grokipedia.com",
    "linkedin.com",
    "reddit.com",
    "twitter.com",
    "x.com",
    "facebook.com",
    "instagram.com",
    "medium.com",
}

SOURCE_COLUMNS = ["P", "Q", "R", "S", "T", "U", "V", "W"]
DATE_COLUMNS = {"U", "V"}
FALLBACK_LITERAL = "fDi Markets (FT)"
UNVERIFIED_PREFIX = "[UNVERIFIED] "
YEAR_ONLY_SUFFIX = "(year only)"


def host_of(url: str) -> str | None:
    try:
        host = urlparse(url).hostname
    except ValueError:
        return None
    if not host:
        return None
    host = host.lower()
    if host.startswith("www."):
        host = host[4:]
    return host


def is_forbidden(url: str) -> str | None:
    host = host_of(url)
    if host is None:
        return None
    for bad in FORBIDDEN_HOSTS:
        if host == bad or host.endswith("." + bad):
            return bad
    return None


def extract_url(cell_value: str) -> str | None:
    """Pull the URL out of a cell that may have prefixes / suffixes."""
    body = cell_value
    if body.startswith(UNVERIFIED_PREFIX):
        body = body[len(UNVERIFIED_PREFIX):]
    body = body.strip()
    if body == FALLBACK_LITERAL:
        return None
    match = re.match(r"(https?://\S+)", body)
    return match.group(1) if match else None


def validate_cell(col: str, value) -> list[str]:
    if value is None:
        return []
    text = str(value).strip()
    if not text:
        return []
    if text == FALLBACK_LITERAL or text == UNVERIFIED_PREFIX + FALLBACK_LITERAL:
        return []

    body = text[len(UNVERIFIED_PREFIX):].strip() if text.startswith(UNVERIFIED_PREFIX) else text

    if col in DATE_COLUMNS and body.endswith(YEAR_ONLY_SUFFIX):
        body = body[: -len(YEAR_ONLY_SUFFIX)].strip()

    if not body.startswith(("http://", "https://")):
        return [f"col {col}: not a URL or recognized fallback: {text!r}"]

    url = extract_url(text)
    if url is None:
        return [f"col {col}: could not parse URL from {text!r}"]

    bad = is_forbidden(url)
    if bad:
        return [f"col {col}: forbidden domain {bad} in {url}"]
    return []


def main(path: str) -> int:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    errors: list[str] = []

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=False))
    col_letter = {cell.column_letter: cell.column_letter for cell in header_row}
    target_cols = [c for c in SOURCE_COLUMNS if c in col_letter]
    if len(target_cols) < len(SOURCE_COLUMNS):
        missing = set(SOURCE_COLUMNS) - set(target_cols)
        print(f"warning: columns missing from sheet: {sorted(missing)}", file=sys.stderr)

    for row in ws.iter_rows(min_row=2, values_only=False):
        for cell in row:
            if cell.column_letter not in SOURCE_COLUMNS:
                continue
            row_errors = validate_cell(cell.column_letter, cell.value)
            for err in row_errors:
                errors.append(f"row {cell.row}, {err}")

    if errors:
        for e in errors:
            print(e)
        print(f"\n{len(errors)} error(s) found in {path}", file=sys.stderr)
        return 1

    print(f"OK: {path} — no forbidden domains or malformed source cells.")
    return 0


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print(__doc__, file=sys.stderr)
        sys.exit(2)
    sys.exit(main(sys.argv[1]))
